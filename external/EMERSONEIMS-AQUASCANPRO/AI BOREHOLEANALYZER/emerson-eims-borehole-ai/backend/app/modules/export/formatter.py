"""
Multi-Format Export Handler
XLSX, CSV, GeoJSON, Shapefile, KML, JSON export capabilities
"""

import json
import csv
from datetime import datetime
from typing import Dict, List, Optional, BinaryIO
import logging

logger = logging.getLogger(__name__)


class ExportFormatter:
    """Handles export of analysis results in multiple formats"""
    
    SUPPORTED_FORMATS = ['xlsx', 'csv', 'geojson', 'shapefile', 'kml', 'json', 'pdf']
    
    def __init__(self):
        self.timestamp = datetime.now().isoformat()
    
    def export_csv(self, analysis_result: Dict) -> str:
        """
        Export analysis to CSV format
        One row per result with all parameters
        """
        try:
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Header
            site = analysis_result.get('site', {})
            soil = analysis_result.get('soil', {})
            water_quality = analysis_result.get('waterQuality', {})
            risk = analysis_result.get('risk', {})
            
            headers = [
                'latitude', 'longitude', 'site_type', 'confidence',
                'probability', 'recommended_depth', 'estimated_yield',
                'soil_type', 'soil_porosity', 'soil_permeability',
                'water_potable', 'tds_mg_l', 'ph', 'fluoride_mg_l',
                'risk_level', 'timestamp'
            ]
            
            writer.writerow(headers)
            
            # Data row
            row = [
                site.get('latitude', ''),
                site.get('longitude', ''),
                site.get('siteType', ''),
                site.get('confidence', ''),
                analysis_result.get('probability', ''),
                analysis_result.get('recommendedDepth', ''),
                analysis_result.get('estimatedYield', ''),
                soil.get('type', ''),
                soil.get('porosity', ''),
                soil.get('permeability', ''),
                water_quality.get('isPotable', ''),
                water_quality.get('tds', ''),
                water_quality.get('pH', ''),
                water_quality.get('fluoride', ''),
                risk.get('overallRisk', ''),
                analysis_result.get('timestamp', '')
            ]
            
            writer.writerow(row)
            
            return output.getvalue()
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return ""
    
    def export_geojson(self, analysis_result: Dict) -> str:
        """
        Export as GeoJSON FeatureCollection
        Includes site location and risk zones
        """
        try:
            site = analysis_result.get('site', {})
            risk = analysis_result.get('risk', {})
            
            features = []
            
            # Site location feature
            site_feature = {
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [site.get('longitude', 0), site.get('latitude', 0)]
                },
                "properties": {
                    "site_type": site.get('siteType'),
                    "confidence": site.get('confidence'),
                    "vegetation_density": site.get('vegetationDensity'),
                    "water_indicator": site.get('waterIndicator'),
                    "terrain_slope": site.get('terrainSlope'),
                    "probability": analysis_result.get('probability'),
                    "depth_m": analysis_result.get('recommendedDepth'),
                    "yield_m3_h": analysis_result.get('estimatedYield')
                }
            }
            features.append(site_feature)
            
            # Risk zones (1 km buffer)
            if risk.get('overallRisk'):
                # Create a square buffer zone
                lat = site.get('latitude', 0)
                lon = site.get('longitude', 0)
                buffer = 0.01  # ~1 km at equator
                
                zone_feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Polygon",
                        "coordinates": [[
                            [lon - buffer, lat - buffer],
                            [lon + buffer, lat - buffer],
                            [lon + buffer, lat + buffer],
                            [lon - buffer, lat + buffer],
                            [lon - buffer, lat - buffer]
                        ]]
                    },
                    "properties": {
                        "feature_type": "risk_zone",
                        "overall_risk": risk.get('overallRisk'),
                        "geological_risk": risk.get('categories', {}).get('geological'),
                        "contamination_risk": risk.get('categories', {}).get('contamination')
                    }
                }
                features.append(zone_feature)
            
            geojson = {
                "type": "FeatureCollection",
                "features": features,
                "properties": {
                    "generated": datetime.now().isoformat(),
                    "crs": "EPSG:4326"
                }
            }
            
            return json.dumps(geojson, indent=2)
        except Exception as e:
            logger.error(f"GeoJSON export failed: {e}")
            return ""
    
    def export_shapefile_sql(self, analysis_result: Dict) -> Dict:
        """
        Generate SQL schema for Shapefile table
        Returns commands to create spatial tables
        """
        try:
            site = analysis_result.get('site', {})
            risk = analysis_result.get('risk', {})
            
            sql_commands = {
                "create_table": """
CREATE TABLE borehole_sites (
    id SERIAL PRIMARY KEY,
    site_name VARCHAR(100),
    latitude DECIMAL(10, 8),
    longitude DECIMAL(11, 8),
    site_type VARCHAR(20),
    confidence DECIMAL(3, 2),
    probability DECIMAL(3, 2),
    recommended_depth INTEGER,
    estimated_yield DECIMAL(5, 2),
    soil_type VARCHAR(50),
    soil_porosity DECIMAL(4, 3),
    overall_risk_level DECIMAL(3, 2),
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    geom GEOMETRY(POINT, 4326)
);

CREATE SPATIAL INDEX idx_borehole_sites_geom ON borehole_sites USING GIST (geom);
                """,
                "insert_statement": f"""
INSERT INTO borehole_sites (
    latitude, longitude, site_type, confidence, probability,
    recommended_depth, estimated_yield, soil_type, soil_porosity,
    overall_risk_level, geom
) VALUES (
    {site.get('latitude', 0)},
    {site.get('longitude', 0)},
    '{site.get('siteType', 'flat')}',
    {site.get('confidence', 0.5)},
    {analysis_result.get('probability', 0.5)},
    {analysis_result.get('recommendedDepth', 50)},
    {analysis_result.get('estimatedYield', 5)},
    '{analysis_result.get('soil', {}).get('type', 'loamy')}',
    {analysis_result.get('soil', {}).get('porosity', 0.2)},
    {risk.get('overallRisk', 0.5)},
    ST_SetSRID(ST_MakePoint({site.get('longitude', 0)}, {site.get('latitude', 0)}), 4326)
);
                """
            }
            
            return sql_commands
        except Exception as e:
            logger.error(f"Shapefile SQL generation failed: {e}")
            return {}
    
    def export_kml(self, analysis_result: Dict) -> str:
        """
        Export as KML for Google Earth visualization
        """
        try:
            site = analysis_result.get('site', {})
            risk = analysis_result.get('risk', {})
            
            kml = f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document>
    <name>Borehole Analysis Results</name>
    <description>AI-generated borehole site analysis</description>
    
    <Style id="siteMarker">
      <IconStyle>
        <Icon>
          <href>http://maps.google.com/mapfiles/kml/shapes/water.png</href>
        </Icon>
      </IconStyle>
      <BalloonStyle>
        <text><![CDATA[
          <h3>Borehole Site Analysis</h3>
          <p><b>Probability:</b> {analysis_result.get('probability', 0):.1%}</p>
          <p><b>Depth:</b> {analysis_result.get('recommendedDepth', 0)} m</p>
          <p><b>Yield:</b> {analysis_result.get('estimatedYield', 0):.1f} m³/h</p>
          <p><b>Risk Level:</b> {risk.get('overallRisk', 0):.2f}</p>
        ]]></text>
      </BalloonStyle>
    </Style>
    
    <Placemark>
      <name>Borehole Site</name>
      <description>Recommended drilling location</description>
      <styleUrl>#siteMarker</styleUrl>
      <Point>
        <coordinates>
          {site.get('longitude', 0)},{site.get('latitude', 0)},0
        </coordinates>
      </Point>
    </Placemark>
    
  </Document>
</kml>"""
            
            return kml
        except Exception as e:
            logger.error(f"KML export failed: {e}")
            return ""
    
    def export_json(self, analysis_result: Dict, pretty: bool = True) -> str:
        """Export as JSON with optional pretty formatting"""
        try:
            if pretty:
                return json.dumps(analysis_result, indent=2, default=str)
            else:
                return json.dumps(analysis_result, default=str)
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            return ""
    
    def export_xlsx(self, analysis_results: List[Dict]) -> bytes:
        """
        Export multiple analysis results to XLSX
        Each result in a separate row with summary statistics
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analysis Results"
            
            # Headers
            headers = [
                'Site ID', 'Latitude', 'Longitude', 'Site Type',
                'Probability (%)', 'Depth (m)', 'Yield (m³/h)',
                'Overall Risk', 'Soil Type', 'Water Potable',
                'TDS (mg/L)', 'Timestamp'
            ]
            
            ws.append(headers)
            
            # Format header row
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for idx, result in enumerate(analysis_results, 1):
                site = result.get('site', {})
                row = [
                    idx,
                    site.get('latitude', ''),
                    site.get('longitude', ''),
                    site.get('siteType', ''),
                    result.get('probability', 0) * 100,
                    result.get('recommendedDepth', ''),
                    result.get('estimatedYield', ''),
                    result.get('risk', {}).get('overallRisk', ''),
                    result.get('soil', {}).get('type', ''),
                    result.get('waterQuality', {}).get('isPotable', ''),
                    result.get('waterQuality', {}).get('tds', ''),
                    result.get('timestamp', '')
                ]
                ws.append(row)
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save to bytes
            import io
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        except ImportError:
            logger.warning("openpyxl not installed. XLSX export unavailable.")
            return b""
        except Exception as e:
            logger.error(f"XLSX export failed: {e}")
            return b""
    
    def export(
        self,
        analysis_result: Dict,
        format: str,
        multiple_results: Optional[List[Dict]] = None
    ) -> str | bytes:
        """
        Main export method
        
        Args:
            analysis_result: Single analysis result
            format: Export format (csv, geojson, kml, json, xlsx)
            multiple_results: Optional list of results for batch export
        
        Returns:
            Exported data as string or bytes
        """
        if format.lower() == 'csv':
            return self.export_csv(analysis_result)
        elif format.lower() == 'geojson':
            return self.export_geojson(analysis_result)
        elif format.lower() == 'kml':
            return self.export_kml(analysis_result)
        elif format.lower() == 'json':
            return self.export_json(analysis_result)
        elif format.lower() == 'xlsx':
            results = multiple_results or [analysis_result]
            return self.export_xlsx(results)
        elif format.lower() == 'shapefile':
            return json.dumps(self.export_shapefile_sql(analysis_result))
        else:
            logger.error(f"Unsupported format: {format}")
            return ""
