"""Tabular export -- CSV (stdlib) and XLSX (openpyxl, optional).

Accepts a JSON payload with one or more 'sheets' and renders a workbook
or a single CSV. Used for QS deliverables (BOQ, cost plans, cashflow,
EVM tables, MC distributions, etc.).

Endpoints:
  POST /api/export/csv    -> single CSV file
  POST /api/export/xlsx   -> multi-sheet XLSX (openpyxl) or JSON fallback
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
import json
import logging
from typing import Any, Dict, List, Optional

logger = logging.getLogger('eims.export')

try:
    from openpyxl import Workbook                       # type: ignore
    from openpyxl.styles import Font, Alignment, PatternFill  # type: ignore
    OPENPYXL_OK = True
except Exception:  # pragma: no cover
    OPENPYXL_OK = False


def render_csv(*, rows: List[Dict[str, Any]],
                  columns: Optional[List[str]] = None) -> bytes:
    if not rows:
        return b''
    cols = columns or sorted({k for r in rows for k in r.keys()})
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=cols, extrasaction='ignore')
    w.writeheader()
    for r in rows:
        w.writerow({c: r.get(c, '') for c in cols})
    return buf.getvalue().encode('utf-8-sig')


def render_xlsx(*, sheets: List[Dict[str, Any]],
                  title: str = 'EmersonEIMS Export') -> bytes:
    """sheets = [{'name', 'columns'?, 'rows', 'meta'?}]"""
    if not OPENPYXL_OK:
        return json.dumps({'_warning': 'openpyxl not installed; JSON fallback',
                            'title': title, 'sheets': sheets},
                            indent=2, default=str).encode('utf-8')

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E78')
    centre = Alignment(horizontal='center')

    for sh in sheets:
        ws = wb.create_sheet(title=str(sh.get('name', 'Sheet'))[:31])
        rows = sh.get('rows') or []
        cols = sh.get('columns') or sorted({k for r in rows for k in r.keys()})
        # header row
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=j, value=str(c))
            cell.font = header_font; cell.fill = header_fill; cell.alignment = centre
        # data
        for i, r in enumerate(rows, 2):
            for j, c in enumerate(cols, 1):
                v = r.get(c, '')
                # native python types only -- avoid openpyxl rejecting dicts
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, default=str)
                ws.cell(row=i, column=j, value=v)
        # autosize-ish
        for j, c in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = \
                min(60, max(12, len(str(c)) + 2))
        # meta block (optional, written below the table)
        meta = sh.get('meta') or {}
        if meta:
            base = len(rows) + 3
            ws.cell(row=base, column=1, value='Source / provenance').font = Font(bold=True)
            for k, (key, val) in enumerate(meta.items(), 1):
                ws.cell(row=base + k, column=1, value=str(key))
                ws.cell(row=base + k, column=2, value=str(val))

    # workbook-level metadata
    wb.properties.title = title
    wb.properties.creator = 'EmersonEIMS Building Suite Pro'
    wb.properties.created = _dt.datetime.now(_dt.timezone.utc)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================================
# Flask integration
# ============================================================================

def register(app, *, auth_required=None) -> None:
    from flask import Response, jsonify, request

    def _csv():
        d = request.get_json(silent=True) or {}
        rows = d.get('rows') or []
        if not isinstance(rows, list):
            return jsonify({'success': False, 'error': 'rows must be a list'}), 400
        try:
            blob = render_csv(rows=rows, columns=d.get('columns'))
        except (TypeError, ValueError) as e:
            return jsonify({'success': False, 'error': str(e)}), 400
        fname = d.get('filename') or f"eims_export_{_dt.datetime.now(_dt.timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
        return Response(blob, mimetype='text/csv',
                          headers={'Content-Disposition': f'attachment; filename={fname}'})

    def _xlsx():
        d = request.get_json(silent=True) or {}
        sheets = d.get('sheets')
        if not sheets or not isinstance(sheets, list):
            return jsonify({'success': False,
                             'error': 'sheets must be a non-empty list'}), 400
        try:
            blob = render_xlsx(sheets=sheets, title=str(d.get('title', 'EmersonEIMS Export')))
        except (TypeError, ValueError) as e:
            return jsonify({'success': False, 'error': str(e)}), 400
        mime = ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                  if OPENPYXL_OK else 'application/json')
        ext = 'xlsx' if OPENPYXL_OK else 'json'
        fname = d.get('filename') or f"eims_export_{_dt.datetime.now(_dt.timezone.utc).strftime('%Y%m%d_%H%M%S')}.{ext}"
        return Response(blob, mimetype=mime,
                          headers={'Content-Disposition': f'attachment; filename={fname}'})

    if auth_required:
        _csv  = auth_required(_csv)
        _xlsx = auth_required(_xlsx)
    app.add_url_rule('/api/export/csv',  'eims_export_csv',  _csv,  methods=['POST'])
    app.add_url_rule('/api/export/xlsx', 'eims_export_xlsx', _xlsx, methods=['POST'])

    logger.info('Export module registered (openpyxl=%s)', OPENPYXL_OK)
